# ==============================================================================
# UNIFIED MTC DATA EXTRACTION SCRIPT WITH EXCEL EXPORT
# ==============================================================================

import zipfile
import xml.etree.ElementTree as ET
import os
import re
import openpyxl  # <--- NEW IMPORT
from pdfminer.high_level import extract_pages
from pdfminer.layout import LTTextContainer

# ==============================================================================
# PART 1: DOCX MICROSTRUCTURE EXTRACTION
# ==============================================================================

def extract_micro_data_from_docx(docx_path):
    """
    Extracts microstructure values from a DOCX file.
    """
    results = {}
    print(f"\n--- X-RAY SCANNING: {docx_path} ---")

    try:
        if not os.path.exists(docx_path):
            print("Error: DOCX File not found.")
            return results
            
        with zipfile.ZipFile(docx_path) as docx:
            xml_content = docx.read('word/document.xml')
    except Exception as e:
        print(f"Error reading DOCX content: {e}")
        return results

    tree = ET.fromstring(xml_content)
    all_text_chunks = []
    for elem in tree.iter():
        if elem.tag.endswith('}t'):
            text_content = elem.text
            if text_content is not None and text_content.strip():
                all_text_chunks.append(text_content.strip())

    target_labels = {
        "Graphite Nodularity": "last",
        "Nodular Particles per mm²": "last",
        "Graphite Size": "last",
        "Graphite Form": "last",
        "Graphite Fraction": "last",
        "Ferrite / Pearlite Ratio": "first"
    }

    print("-" * 40)
    
    for label, occurrence_preference in target_labels.items():
        found_value = None # Changed default to None for easier checking later
        target_index = -1
        
        if occurrence_preference == "last":
            for i, chunk in enumerate(all_text_chunks):
                if label.lower() in chunk.lower():
                    target_index = i
        elif occurrence_preference == "first":
            for i, chunk in enumerate(all_text_chunks):
                if label.lower() in chunk.lower():
                    target_index = i
                    break

        if target_index != -1:
            neighbors = all_text_chunks[target_index+1:target_index+6]
            
            if label == "Graphite Fraction":
                for j, n in enumerate(neighbors):
                    if "%" in n and any(c.isdigit() for c in n):
                        found_value = n
                        break
                    if any(c.isdigit() for c in n) and j + 1 < len(neighbors) and neighbors[j+1] == "%":
                        found_value = f"{n}{neighbors[j+1]}"
                        break
            
            elif label == "Graphite Form":
                for n in neighbors:
                     if ("(" in n and ")" in n):
                        found_value = n
                        break

            elif label == "Ferrite / Pearlite Ratio":
                combined = "".join(neighbors[0:3])
                match = re.search(r"(\d+\.?\d*%\s*/\s*\d+\.?\d*%)", combined)
                if match:
                    found_value = match.group(1)
            
            elif label == "Graphite Nodularity":
                for n in neighbors:
                    if "%" in n and len(n) > 1:
                        found_value = n
                        break
            
            elif label in ["Nodular Particles per mm²", "Graphite Size"]:
                for n in neighbors:
                    if any(c.isdigit() for c in n) and not n.endswith('%'):
                        found_value = re.sub(r'[\s\.\,]+$', '', n)
                        break
            
            if found_value:
                results[label] = found_value
                print(f"{label:30} : '{found_value}'")
            else:
                results[label] = "Not Found"
                print(f"{label:30} : Not Found")
            
    print("-" * 40)
    return results

# ==============================================================================
# PART 2: PDF MECHANICAL ANALYSIS EXTRACTION
# ==============================================================================

def find_value_neighbor(elements, label_text, required_keyword="Mpa"):
    label_bbox = None
    for element in elements:
        if label_text in element.get_text():
            label_bbox = element.bbox
            break    
    if not label_bbox: return "Label Not Found"

    lx0, ly0, lx1, ly1 = label_bbox
    best_candidate_text = None
    closest_distance = 9999
    
    for element in elements:
        text = element.get_text().strip()
        ex0, ey0, ex1, ey1 = element.bbox
        if label_text in text: continue
        
        is_vertically_aligned = (ey0 < ly1 + 2) and (ey1 > ly0 - 2)
        is_to_the_right = ex0 >= (lx0 - 5)
        has_keyword = required_keyword in text
        
        if is_vertically_aligned and is_to_the_right and has_keyword:
            distance = ex0 - lx1
            if distance < closest_distance:
                closest_distance = distance
                best_candidate_text = text
    return best_candidate_text

def extract_number_only(text):
    if not text: return None
    match = re.search(r"([\d\.]+)", text)
    if match: return match.group(1)
    return text

def process_tensile_file(pdf_path):
    if not os.path.exists(pdf_path):
        print(f"Error: PDF File not found at {pdf_path}")
        return None, None, None
        
    print(f"\n--- Processing Tensile Report: {pdf_path} ---")
    elements = []
    for page_layout in extract_pages(pdf_path, page_numbers=[0]):
        for element in page_layout:
            if isinstance(element, LTTextContainer):
                elements.append(element)

    raw_tensile = find_value_neighbor(elements, "Tensile Strength", "Mpa")
    val_tensile = extract_number_only(raw_tensile)
    
    raw_yield = find_value_neighbor(elements, "Yield Strength", "Mpa")
    val_yield = extract_number_only(raw_yield)

    raw_elongation = find_value_neighbor(elements, "Elongation", "%")
    val_elongation = extract_number_only(raw_elongation)

    print(f"Tensile Strength: '{val_tensile}'")
    print(f"Yield Strength:   '{val_yield}'")
    print(f"Elongation:       '{val_elongation}'")
    return val_tensile, val_yield, val_elongation

def process_hardness_file(pdf_path):
    if not os.path.exists(pdf_path):
        print(f"Error: PDF File not found at {pdf_path}")
        return []
        
    print(f"\n--- Processing Hardness Report: {pdf_path} ---")
    
    elements = []
    for page_layout in extract_pages(pdf_path, page_numbers=[0]):
        for element in page_layout:
            if isinstance(element, LTTextContainer):
                elements.append(element)

    hardness_labels = [e for e in elements if "Hardness" in e.get_text()]
    hardness_labels.sort(key=lambda x: x.bbox[3], reverse=True)
    
    extracted_values = []
    count = 1

    for label in hardness_labels:
        lx0, ly0, lx1, ly1 = label.bbox
        label_text = label.get_text().strip()
        found_val = None

        match_inside = re.search(r"([\d\.]+)\s*HBW", label_text)
        if match_inside:
            found_val = match_inside.group(1)
        
        if not found_val:
            closest_dist = 9999
            for element in elements:
                etext = element.get_text().strip()
                ex0, ey0, ex1, ey1 = element.bbox
                if "HBW" not in etext: continue
                if (ey0 < ly1 + 5) and (ey1 > ly0 - 5):
                    if ex0 > lx0:
                        dist = ex0 - lx1
                        if dist < closest_dist:
                            n_match = re.search(r"([\d\.]+)\s*HBW", etext)
                            if n_match:
                                closest_dist = dist
                                found_val = n_match.group(1)

        if found_val:
            print(f"({count}) Found Hardness: '{found_val}'")
            extracted_values.append(found_val)
            count += 1
            
    return extracted_values

# ==============================================================================
# PART 3: EXCEL WRITING (NEW ADDITION)
# ==============================================================================

def update_excel_mtc(excel_path, micro_data, tensile_data, hardness_data):
    """
    Writes the extracted data into the specific Excel cells.
    """
    print(f"\n--- Writing to Excel: {excel_path} ---")

    if not os.path.exists(excel_path):
        print("CRITICAL ERROR: Excel file does not exist!")
        return

    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_path)
        
        # Select active sheet (or specify name like wb['Sheet1'])
        ws = wb.active 

        # --- 1. Tensile Data (Tuple: Tensile, Yield, Elongation) ---
        # ultimate tencile strength = e26
        # yield strength = e27
        # elongation = e28
        val_tensile, val_yield, val_elongation = tensile_data
        
        if val_tensile: ws['E26'] = val_tensile
        if val_yield:   ws['E27'] = val_yield
        if val_elongation: ws['E28'] = val_elongation

        # --- 2. Hardness Data (List) ---
        # hardness(BHN) = e29 and e30 (separately)
        if len(hardness_data) > 0:
            ws['E29'] = hardness_data[0]
        if len(hardness_data) > 1:
            ws['E30'] = hardness_data[1]

        # --- 3. Microstructure Data (Dictionary) ---
        # Note: openpyxl writes to the top-left cell of a merged range.
        # So for t38:u38, we just write to T38.
        
        # graphite nodularity by count = t36:u36 -> T36
        if "Graphite Nodularity" in micro_data:
            ws['T36'] = micro_data["Graphite Nodularity"]

        # nodularity particle per mm = t37:u37 -> T37
        if "Nodular Particles per mm²" in micro_data:
            ws['T37'] = micro_data["Nodular Particles per mm²"]

        # Graphite Size = t38:u38 -> T38
        if "Graphite Size" in micro_data:
            ws['T38'] = micro_data["Graphite Size"]

        # graphite form = t39:u39 -> T39
        if "Graphite Form" in micro_data:
            ws['T39'] = micro_data["Graphite Form"]

        # graphite fraction = t40:u40 -> T40
        if "Graphite Fraction" in micro_data:
            ws['T40'] = micro_data["Graphite Fraction"]

        # ferrite / pearlite ratio = t41:u41 -> T41
        if "Ferrite / Pearlite Ratio" in micro_data:
            ws['T41'] = micro_data["Ferrite / Pearlite Ratio"]

        # Save the file
        wb.save(excel_path)
        print("Success! Data successfully saved to Excel.")

    except PermissionError:
        print("ERROR: Permission denied. Please CLOSE the Excel file and try again.")
    except Exception as e:
        print(f"An error occurred while writing to Excel: {e}")

# ==============================================================================
# MAIN EXECUTION BLOCK (Unified Entry Point)
# ==============================================================================

if __name__ == "__main__":
    # --- DEFINE FILE PATHS HERE ---
    path_micro_report = "/home/johnny/MTCAUTO/MTCAUTO/MICRO_REPORT/F305-013-(BE406)-6.docx"
    path_tensile_report = "/home/johnny/MTCAUTO/MTCAUTO/TENCILE_OG/F326-029(AF427)-4(9).pdf"
    path_hardness_report = "/home/johnny/MTCAUTO/MTCAUTO/HARDNESS_OG/F335-023(AF577)-2_ON_CASTING.pdf"
    
    # Target Excel File
    path_excel_output = "/home/johnny/MTCAUTO/MTCAUTO/play_MTC.xlsx"

    # 1. PROCESS MICROSTRUCTURE REPORT (DOCX)
    micro_data = extract_micro_data_from_docx(path_micro_report)

    # 2. PROCESS TENSILE REPORT (PDF)
    tensile_data = process_tensile_file(path_tensile_report)

    # 3. PROCESS HARDNESS REPORT (PDF)
    hardness_data = process_hardness_file(path_hardness_report)
    
    # 4. WRITE TO EXCEL (NEW STEP)
    update_excel_mtc(path_excel_output, micro_data, tensile_data, hardness_data)
    
    print("\n\n=== ALL JOBS DONE ===")
