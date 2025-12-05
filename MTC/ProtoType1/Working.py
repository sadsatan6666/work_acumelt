import tkinter as tk
from tkinter import filedialog, ttk, messagebox, scrolledtext
import threading
import os
import re
import zipfile
import xml.etree.ElementTree as ET
import openpyxl
from pdfminer.high_level import extract_pages
from pdfminer.layout import LTTextContainer

# ==============================================================================
# PART 1: BACKEND LOGIC (Updated with Logger Callbacks)
# ==============================================================================

def extract_micro_data_from_docx(docx_path, logger=None):
    """ Extracts microstructure values from a DOCX file. """
    def log(msg):
        if logger: logger(msg)

    results = {}
    log(f"--- Scanning DOCX: {os.path.basename(docx_path)} ---")

    try:
        if not os.path.exists(docx_path): 
            log("Error: DOCX File not found.")
            return results
        with zipfile.ZipFile(docx_path) as docx:
            xml_content = docx.read('word/document.xml')
    except Exception as e:
        log(f"Error reading DOCX: {e}")
        return results

    tree = ET.fromstring(xml_content)
    all_text_chunks = []
    for elem in tree.iter():
        if elem.tag.endswith('}t'):
            if elem.text and elem.text.strip():
                all_text_chunks.append(elem.text.strip())

    target_labels = {
        "Graphite Nodularity": "last", "Nodular Particles per mm²": "last",
        "Graphite Size": "last", "Graphite Form": "last",
        "Graphite Fraction": "last", "Ferrite / Pearlite Ratio": "first"
    }
    
    for label, preference in target_labels.items():
        target_index = -1
        found_value = None
        
        # Search strategy
        if preference == "last":
            for i, chunk in enumerate(all_text_chunks):
                if label.lower() in chunk.lower(): target_index = i
        elif preference == "first":
            for i, chunk in enumerate(all_text_chunks):
                if label.lower() in chunk.lower(): 
                    target_index = i
                    break

        if target_index != -1:
            neighbors = all_text_chunks[target_index+1:target_index+6]
            if label == "Graphite Fraction":
                for j, n in enumerate(neighbors):
                    if "%" in n and any(c.isdigit() for c in n):
                        found_value = n; break
                    if any(c.isdigit() for c in n) and j+1 < len(neighbors) and neighbors[j+1] == "%":
                        found_value = f"{n}{neighbors[j+1]}"; break
            elif label == "Graphite Form":
                for n in neighbors:
                     if "(" in n and ")" in n: found_value = n; break
            elif label == "Ferrite / Pearlite Ratio":
                combined = "".join(neighbors[0:3])
                match = re.search(r"(\d+\.?\d*%\s*/\s*\d+\.?\d*%)", combined)
                if match: found_value = match.group(1)
            elif label == "Graphite Nodularity":
                for n in neighbors:
                    if "%" in n and len(n) > 1: found_value = n; break
            elif label in ["Nodular Particles per mm²", "Graphite Size"]:
                for n in neighbors:
                    if any(c.isdigit() for c in n) and not n.endswith('%'):
                        found_value = re.sub(r'[\s\.\,]+$', '', n); break
            
            if found_value: 
                results[label] = found_value
                log(f"[FOUND] {label}: {found_value}")
            else:
                log(f"[MISSING] {label}")
                
    return results

def find_value_neighbor(elements, label_text, required_keyword="Mpa"):
    label_bbox = None
    for element in elements:
        if label_text in element.get_text():
            label_bbox = element.bbox; break    
    if not label_bbox: return "Label Not Found"

    lx0, ly0, lx1, ly1 = label_bbox
    best_candidate_text = None
    closest_distance = 9999
    
    for element in elements:
        text = element.get_text().strip()
        ex0, ey0, ex1, ey1 = element.bbox
        if label_text in text: continue
        
        if (ey0 < ly1 + 2) and (ey1 > ly0 - 2) and (ex0 >= lx0 - 5) and (required_keyword in text):
            distance = ex0 - lx1
            if distance < closest_distance:
                closest_distance = distance
                best_candidate_text = text
    return best_candidate_text

def extract_number_only(text):
    if not text: return None
    match = re.search(r"([\d\.]+)", text)
    if match: return match.group(1)
    return text

def process_tensile_file(pdf_path, logger=None):
    def log(msg):
        if logger: logger(msg)

    log(f"--- Scanning Tensile PDF: {os.path.basename(pdf_path)} ---")
    if not os.path.exists(pdf_path): return None, None, None
    elements = []
    try:
        for page_layout in extract_pages(pdf_path, page_numbers=[0]):
            for element in page_layout:
                if isinstance(element, LTTextContainer): elements.append(element)
    except Exception as e:
        log(f"Error reading Tensile PDF: {e}")
        return None, None, None

    val_tensile = extract_number_only(find_value_neighbor(elements, "Tensile Strength", "Mpa"))
    val_yield = extract_number_only(find_value_neighbor(elements, "Yield Strength", "Mpa"))
    val_elongation = extract_number_only(find_value_neighbor(elements, "Elongation", "%"))
    
    log(f"Tensile: {val_tensile}")
    log(f"Yield: {val_yield}")
    log(f"Elongation: {val_elongation}")
    
    return val_tensile, val_yield, val_elongation

def process_hardness_file(pdf_path, logger=None):
    def log(msg):
        if logger: logger(msg)

    log(f"--- Scanning Hardness PDF: {os.path.basename(pdf_path)} ---")
    if not os.path.exists(pdf_path): return []
    elements = []
    try:
        for page_layout in extract_pages(pdf_path, page_numbers=[0]):
            for element in page_layout:
                if isinstance(element, LTTextContainer): elements.append(element)
    except Exception as e:
        log(f"Error reading Hardness PDF: {e}")
        return []

    hardness_labels = [e for e in elements if "Hardness" in e.get_text()]
    hardness_labels.sort(key=lambda x: x.bbox[3], reverse=True)
    
    extracted_values = []
    count = 1
    for label in hardness_labels:
        lx0, ly0, lx1, ly1 = label.bbox
        label_text = label.get_text().strip()
        found_val = None

        match_inside = re.search(r"([\d\.]+)\s*HBW", label_text)
        if match_inside: found_val = match_inside.group(1)
        
        if not found_val:
            closest_dist = 9999
            for element in elements:
                etext = element.get_text().strip()
                ex0, ey0, ex1, ey1 = element.bbox
                if "HBW" not in etext: continue
                if (ey0 < ly1 + 5) and (ey1 > ly0 - 5) and (ex0 > lx0):
                    dist = ex0 - lx1
                    if dist < closest_dist:
                        n_match = re.search(r"([\d\.]+)\s*HBW", etext)
                        if n_match:
                            closest_dist = dist
                            found_val = n_match.group(1)
        if found_val: 
            extracted_values.append(found_val)
            log(f"Hardness #{count}: {found_val}")
            count += 1
    return extracted_values

def update_excel_mtc(excel_path, micro_data, tensile_data, hardness_data, logger=None, progress_callback=None):
    def log(msg):
        if logger: logger(msg)
    
    def update_prog(val):
        if progress_callback: progress_callback(val)

    log(f"Opening Excel: {os.path.basename(excel_path)}")
    update_prog(10)

    if not os.path.exists(excel_path): raise FileNotFoundError("Excel file not found")
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active 

    # Calculate Total Operations for Progress Bar
    total_ops = 3 + 2 + 6 + 1 # Tensile + Hardness + Micro + Save
    current_op = 0

    def step(desc):
        nonlocal current_op
        current_op += 1
        percentage = 10 + (current_op / total_ops * 80) # Scale between 10% and 90%
        update_prog(percentage)
        log(desc)

    # 1. Tensile
    val_tensile, val_yield, val_elongation = tensile_data
    if val_tensile: 
        ws['E26'] = val_tensile
        step(f"Set E26 = {val_tensile} (Tensile)")
    if val_yield:   
        ws['E27'] = val_yield
        step(f"Set E27 = {val_yield} (Yield)")
    if val_elongation: 
        ws['E28'] = val_elongation
        step(f"Set E28 = {val_elongation} (Elong)")

    # 2. Hardness
    if len(hardness_data) > 0: 
        ws['E29'] = hardness_data[0]
        step(f"Set E29 = {hardness_data[0]} (BHN 1)")
    if len(hardness_data) > 1: 
        ws['E30'] = hardness_data[1]
        step(f"Set E30 = {hardness_data[1]} (BHN 2)")

    # 3. Micro
    mapping = {
        "Graphite Nodularity": 'T36', "Nodular Particles per mm²": 'T37',
        "Graphite Size": 'T38', "Graphite Form": 'T39',
        "Graphite Fraction": 'T40', "Ferrite / Pearlite Ratio": 'T41'
    }
    for key, cell in mapping.items():
        if key in micro_data: 
            ws[cell] = micro_data[key]
            step(f"Set {cell} = {micro_data[key]} ({key})")

    log("Saving Excel file...")
    wb.save(excel_path)
    update_prog(100)
    log("Excel Saved Successfully.")

# ==============================================================================
# PART 2: THE UI (TKINTER)
# ==============================================================================

class MTCApp:
    def __init__(self, root):
        self.root = root
        self.root.title("MTC Automation Tool")
        self.root.geometry("900x700") # Increased size for logs
        self.root.resizable(True, True)

        # Variables
        self.path_micro = tk.StringVar()
        self.path_tensile = tk.StringVar()
        self.path_hardness = tk.StringVar()
        self.path_excel = tk.StringVar()

        self.create_widgets()

    def create_widgets(self):
        # --- Header ---
        header = tk.Label(self.root, text="MTC Data Extraction & Filler", font=("Helvetica", 16, "bold"))
        header.pack(pady=10)

        # --- File Inputs ---
        frame_files = tk.LabelFrame(self.root, text="File Selection", padx=10, pady=10)
        frame_files.pack(padx=20, pady=5, fill="x")

        self.create_file_row(frame_files, "Micro Report (.docx):", self.path_micro, [("Word files", "*.docx")])
        self.create_file_row(frame_files, "Tensile Report (.pdf):", self.path_tensile, [("PDF files", "*.pdf")])
        self.create_file_row(frame_files, "Hardness Report (.pdf):", self.path_hardness, [("PDF files", "*.pdf")])
        ttk.Separator(frame_files, orient='horizontal').pack(fill='x', pady=5)
        self.create_file_row(frame_files, "MTC Excel File (.xlsx):", self.path_excel, [("Excel files", "*.xlsx")])

        # --- Logs Section (Split View) ---
        frame_logs = tk.Frame(self.root)
        frame_logs.pack(padx=20, pady=10, fill="both", expand=True)

        # Left Log: Extracted Data
        frame_left = tk.Frame(frame_logs)
        frame_left.pack(side="left", fill="both", expand=True, padx=(0, 5))
        tk.Label(frame_left, text="Fetched Values Log", font=("Arial", 10, "bold"), fg="blue").pack(anchor="w")
        self.log_data_widget = scrolledtext.ScrolledText(frame_left, height=10, state='disabled', font=("Consolas", 9))
        self.log_data_widget.pack(fill="both", expand=True)

        # Right Log: Writing Process
        frame_right = tk.Frame(frame_logs)
        frame_right.pack(side="right", fill="both", expand=True, padx=(5, 0))
        tk.Label(frame_right, text="Data Writing Log", font=("Arial", 10, "bold"), fg="green").pack(anchor="w")
        self.log_write_widget = scrolledtext.ScrolledText(frame_right, height=10, state='disabled', font=("Consolas", 9))
        self.log_write_widget.pack(fill="both", expand=True)

        # --- Progress Section ---
        frame_prog = tk.Frame(self.root)
        frame_prog.pack(padx=20, pady=10, fill="x")

        # Main Progress
        tk.Label(frame_prog, text="Overall Progress:").pack(anchor="w")
        self.progress_main = ttk.Progressbar(frame_prog, orient="horizontal", length=500, mode="determinate")
        self.progress_main.pack(fill="x", pady=(0, 10))

        # Write Progress
        tk.Label(frame_prog, text="Data Writing Progress:").pack(anchor="w")
        self.progress_write = ttk.Progressbar(frame_prog, orient="horizontal", length=500, mode="determinate")
        self.progress_write.pack(fill="x")
        
        # Status Label
        self.status_label = tk.Label(self.root, text="Ready", fg="gray")
        self.status_label.pack()

        # Run Button
        self.btn_run = tk.Button(self.root, text="START EXTRACTION", command=self.start_thread, 
                                 bg="#4CAF50", fg="white", font=("Arial", 12, "bold"), height=2, width=20)
        self.btn_run.pack(pady=10)

    def create_file_row(self, parent, label_text, variable, file_types):
        row_frame = tk.Frame(parent)
        row_frame.pack(fill="x", pady=2)
        tk.Label(row_frame, text=label_text, width=20, anchor="w").pack(side="left")
        tk.Entry(row_frame, textvariable=variable, width=50).pack(side="left", padx=5, fill="x", expand=True)
        tk.Button(row_frame, text="Browse", command=lambda: self.browse_file(variable, file_types)).pack(side="left")

    def browse_file(self, variable, file_types):
        filename = filedialog.askopenfilename(filetypes=file_types)
        if filename: variable.set(filename)

    # --- Logger Helpers ---
    def log_data(self, message):
        self.root.after(0, lambda: self._append_log(self.log_data_widget, message))

    def log_write(self, message):
        self.root.after(0, lambda: self._append_log(self.log_write_widget, message))

    def _append_log(self, widget, message):
        widget.config(state='normal')
        widget.insert(tk.END, message + "\n")
        widget.see(tk.END) # Auto scroll
        widget.config(state='disabled')

    def update_write_progress(self, value):
        self.root.after(0, lambda: self.progress_write.configure(value=value))

    # --- Threading & Execution ---
    def start_thread(self):
        if not all([self.path_micro.get(), self.path_tensile.get(), self.path_hardness.get(), self.path_excel.get()]):
            messagebox.showwarning("Missing Files", "Please select all 4 files before starting.")
            return

        self.btn_run.config(state="disabled", text="Processing...")
        
        # Clear logs
        self.log_data_widget.config(state='normal'); self.log_data_widget.delete(1.0, tk.END); self.log_data_widget.config(state='disabled')
        self.log_write_widget.config(state='normal'); self.log_write_widget.delete(1.0, tk.END); self.log_write_widget.config(state='disabled')
        self.progress_main['value'] = 0
        self.progress_write['value'] = 0

        threading.Thread(target=self.run_process, daemon=True).start()

    def run_process(self):
        try:
            # 1. Micro
            self.update_status("Reading Microstructure...", 5)
            micro_data = extract_micro_data_from_docx(self.path_micro.get(), logger=self.log_data)
            
            # 2. Tensile
            self.update_status("Reading Tensile...", 30)
            tensile_data = process_tensile_file(self.path_tensile.get(), logger=self.log_data)
            
            # 3. Hardness
            self.update_status("Reading Hardness...", 60)
            hardness_data = process_hardness_file(self.path_hardness.get(), logger=self.log_data)
            
            # 4. Writing
            self.update_status("Writing to Excel...", 85)
            update_excel_mtc(
                self.path_excel.get(), 
                micro_data, 
                tensile_data, 
                hardness_data, 
                logger=self.log_write, 
                progress_callback=self.update_write_progress
            )
            
            self.update_status("Completed!", 100)
            messagebox.showinfo("Success", "Process Completed Successfully!")
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
            self.log_data(f"ERROR: {str(e)}")
            self.update_status("Error", 0)
        finally:
            self.root.after(0, lambda: self.btn_run.config(state="normal", text="START EXTRACTION"))

    def update_status(self, text, progress_val):
        self.root.after(0, lambda: self.status_label.config(text=text))
        self.root.after(0, lambda: self.progress_main.configure(value=progress_val))

# ==============================================================================
# MAIN ENTRY POINT
# ==============================================================================

if __name__ == "__main__":
    root = tk.Tk()
    
    app = MTCApp(root)
    root.mainloop()
