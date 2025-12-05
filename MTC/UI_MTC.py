import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import threading
import os
import re
import zipfile
import xml.etree.ElementTree as ET
import openpyxl
from pdfminer.high_level import extract_pages
from pdfminer.layout import LTTextContainer

# ==============================================================================
# PART 1: BACKEND LOGIC (Your Existing Extraction Code)
# ==============================================================================

def extract_micro_data_from_docx(docx_path):
    results = {}
    try:
        if not os.path.exists(docx_path): return results
        with zipfile.ZipFile(docx_path) as docx:
            xml_content = docx.read('word/document.xml')
    except Exception as e:
        print(f"Error reading DOCX: {e}")
        return results

    tree = ET.fromstring(xml_content)
    all_text_chunks = []
    for elem in tree.iter():
        if elem.tag.endswith('}t'):
            if elem.text and elem.text.strip():
                all_text_chunks.append(elem.text.strip())

    target_labels = {
        "Graphite Nodularity": "last", "Nodular Particles per mm²": "last",
        "Graphite Size": "last", "Graphite Form": "last",
        "Graphite Fraction": "last", "Ferrite / Pearlite Ratio": "first"
    }
    
    for label, preference in target_labels.items():
        target_index = -1
        found_value = None
        
        # Search strategy
        if preference == "last":
            for i, chunk in enumerate(all_text_chunks):
                if label.lower() in chunk.lower(): target_index = i
        elif preference == "first":
            for i, chunk in enumerate(all_text_chunks):
                if label.lower() in chunk.lower(): 
                    target_index = i
                    break

        if target_index != -1:
            neighbors = all_text_chunks[target_index+1:target_index+6]
            # Extraction Logic
            if label == "Graphite Fraction":
                for j, n in enumerate(neighbors):
                    if "%" in n and any(c.isdigit() for c in n):
                        found_value = n; break
                    if any(c.isdigit() for c in n) and j+1 < len(neighbors) and neighbors[j+1] == "%":
                        found_value = f"{n}{neighbors[j+1]}"; break
            elif label == "Graphite Form":
                for n in neighbors:
                     if "(" in n and ")" in n: found_value = n; break
            elif label == "Ferrite / Pearlite Ratio":
                combined = "".join(neighbors[0:3])
                match = re.search(r"(\d+\.?\d*%\s*/\s*\d+\.?\d*%)", combined)
                if match: found_value = match.group(1)
            elif label == "Graphite Nodularity":
                for n in neighbors:
                    if "%" in n and len(n) > 1: found_value = n; break
            elif label in ["Nodular Particles per mm²", "Graphite Size"]:
                for n in neighbors:
                    if any(c.isdigit() for c in n) and not n.endswith('%'):
                        found_value = re.sub(r'[\s\.\,]+$', '', n); break
            
            if found_value: results[label] = found_value
    return results

def find_value_neighbor(elements, label_text, required_keyword="Mpa"):
    label_bbox = None
    for element in elements:
        if label_text in element.get_text():
            label_bbox = element.bbox; break    
    if not label_bbox: return "Label Not Found"

    lx0, ly0, lx1, ly1 = label_bbox
    best_candidate_text = None
    closest_distance = 9999
    
    for element in elements:
        text = element.get_text().strip()
        ex0, ey0, ex1, ey1 = element.bbox
        if label_text in text: continue
        
        if (ey0 < ly1 + 2) and (ey1 > ly0 - 2) and (ex0 >= lx0 - 5) and (required_keyword in text):
            distance = ex0 - lx1
            if distance < closest_distance:
                closest_distance = distance
                best_candidate_text = text
    return best_candidate_text

def extract_number_only(text):
    if not text: return None
    match = re.search(r"([\d\.]+)", text)
    if match: return match.group(1)
    return text

def process_tensile_file(pdf_path):
    if not os.path.exists(pdf_path): return None, None, None
    elements = []
    try:
        for page_layout in extract_pages(pdf_path, page_numbers=[0]):
            for element in page_layout:
                if isinstance(element, LTTextContainer): elements.append(element)
    except Exception as e:
        print(f"Error reading Tensile PDF: {e}")
        return None, None, None

    val_tensile = extract_number_only(find_value_neighbor(elements, "Tensile Strength", "Mpa"))
    val_yield = extract_number_only(find_value_neighbor(elements, "Yield Strength", "Mpa"))
    val_elongation = extract_number_only(find_value_neighbor(elements, "Elongation", "%"))
    return val_tensile, val_yield, val_elongation

def process_hardness_file(pdf_path):
    if not os.path.exists(pdf_path): return []
    elements = []
    try:
        for page_layout in extract_pages(pdf_path, page_numbers=[0]):
            for element in page_layout:
                if isinstance(element, LTTextContainer): elements.append(element)
    except Exception as e:
        print(f"Error reading Hardness PDF: {e}")
        return []

    hardness_labels = [e for e in elements if "Hardness" in e.get_text()]
    hardness_labels.sort(key=lambda x: x.bbox[3], reverse=True)
    
    extracted_values = []
    for label in hardness_labels:
        lx0, ly0, lx1, ly1 = label.bbox
        label_text = label.get_text().strip()
        found_val = None

        match_inside = re.search(r"([\d\.]+)\s*HBW", label_text)
        if match_inside: found_val = match_inside.group(1)
        
        if not found_val:
            closest_dist = 9999
            for element in elements:
                etext = element.get_text().strip()
                ex0, ey0, ex1, ey1 = element.bbox
                if "HBW" not in etext: continue
                if (ey0 < ly1 + 5) and (ey1 > ly0 - 5) and (ex0 > lx0):
                    dist = ex0 - lx1
                    if dist < closest_dist:
                        n_match = re.search(r"([\d\.]+)\s*HBW", etext)
                        if n_match:
                            closest_dist = dist
                            found_val = n_match.group(1)
        if found_val: extracted_values.append(found_val)
    return extracted_values

def update_excel_mtc(excel_path, micro_data, tensile_data, hardness_data):
    if not os.path.exists(excel_path): raise FileNotFoundError("Excel file not found")
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active 

    val_tensile, val_yield, val_elongation = tensile_data
    if val_tensile: ws['E26'] = val_tensile
    if val_yield:   ws['E27'] = val_yield
    if val_elongation: ws['E28'] = val_elongation

    if len(hardness_data) > 0: ws['E29'] = hardness_data[0]
    if len(hardness_data) > 1: ws['E30'] = hardness_data[1]

    mapping = {
        "Graphite Nodularity": 'T36', "Nodular Particles per mm²": 'T37',
        "Graphite Size": 'T38', "Graphite Form": 'T39',
        "Graphite Fraction": 'T40', "Ferrite / Pearlite Ratio": 'T41'
    }
    for key, cell in mapping.items():
        if key in micro_data: ws[cell] = micro_data[key]

    wb.save(excel_path)

# ==============================================================================
# PART 2: THE UI (TKINTER)
# ==============================================================================

class MTCApp:
    def __init__(self, root):
        self.root = root
        self.root.title("MTC Automation Tool")
        self.root.geometry("600x450")
        self.root.resizable(False, False)

        # Variables to store file paths
        self.path_micro = tk.StringVar()
        self.path_tensile = tk.StringVar()
        self.path_hardness = tk.StringVar()
        self.path_excel = tk.StringVar()

        # Build UI
        self.create_widgets()

    def create_widgets(self):
        # Header
        header = tk.Label(self.root, text="MTC Data Extraction & Filler", font=("Helvetica", 16, "bold"))
        header.pack(pady=15)

        # File Selection Frame
        frame = tk.Frame(self.root)
        frame.pack(padx=20, pady=10, fill="x")

        # Row 1: Micro Report
        self.create_file_row(frame, "Micro Report (.docx):", self.path_micro, [("Word files", "*.docx")])
        
        # Row 2: Tensile Report
        self.create_file_row(frame, "Tensile Report (.pdf):", self.path_tensile, [("PDF files", "*.pdf")])

        # Row 3: Hardness Report
        self.create_file_row(frame, "Hardness Report (.pdf):", self.path_hardness, [("PDF files", "*.pdf")])

        # Divider
        ttk.Separator(self.root, orient='horizontal').pack(fill='x', padx=20, pady=15)

        # Row 4: Target Excel
        frame_excel = tk.Frame(self.root)
        frame_excel.pack(padx=20, fill="x")
        self.create_file_row(frame_excel, "MTC Excel File (.xlsx):", self.path_excel, [("Excel files", "*.xlsx")])

        # Progress Bar
        self.progress = ttk.Progressbar(self.root, orient="horizontal", length=500, mode="determinate")
        self.progress.pack(pady=20)
        
        # Status Label
        self.status_label = tk.Label(self.root, text="Ready", fg="gray")
        self.status_label.pack()

        # Run Button
        self.btn_run = tk.Button(self.root, text="START EXTRACTION", command=self.start_thread, 
                                 bg="#4CAF50", fg="white", font=("Arial", 12, "bold"), height=2, width=20)
        self.btn_run.pack(pady=10)

    def create_file_row(self, parent, label_text, variable, file_types):
        """Helper to create a label, entry, and browse button."""
        row_frame = tk.Frame(parent)
        row_frame.pack(fill="x", pady=5)
        
        lbl = tk.Label(row_frame, text=label_text, width=20, anchor="w")
        lbl.pack(side="left")
        
        entry = tk.Entry(row_frame, textvariable=variable, width=40, fg="blue")
        entry.pack(side="left", padx=5)
        
        btn = tk.Button(row_frame, text="Browse", command=lambda: self.browse_file(variable, file_types))
        btn.pack(side="left")

    def browse_file(self, variable, file_types):
        filename = filedialog.askopenfilename(filetypes=file_types)
        if filename:
            variable.set(filename)

    def start_thread(self):
        # Validation
        if not all([self.path_micro.get(), self.path_tensile.get(), self.path_hardness.get(), self.path_excel.get()]):
            messagebox.showwarning("Missing Files", "Please select all 4 files before starting.")
            return

        # Disable button
        self.btn_run.config(state="disabled", text="Processing...")
        
        # Start threading to keep UI responsive
        threading.Thread(target=self.run_process, daemon=True).start()

    def run_process(self):
        try:
            # 0%
            self.update_status("Reading Microstructure Report...", 5)
            micro_data = extract_micro_data_from_docx(self.path_micro.get())
            
            # 30%
            self.update_status("Reading Tensile Report...", 30)
            tensile_data = process_tensile_file(self.path_tensile.get())
            
            # 60%
            self.update_status("Reading Hardness Report...", 60)
            hardness_data = process_hardness_file(self.path_hardness.get())
            
            # 85%
            self.update_status("Writing to Excel...", 85)
            update_excel_mtc(self.path_excel.get(), micro_data, tensile_data, hardness_data)
            
            # 100%
            self.update_status("Completed!", 100)
            messagebox.showinfo("Success", "Data extracted and saved successfully!")
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
            self.update_status("Error", 0)
        finally:
            self.root.after(0, self.reset_ui)

    def update_status(self, text, progress_val):
        # Use root.after to safely update UI from thread
        self.root.after(0, lambda: self.status_label.config(text=text))
        self.root.after(0, lambda: self.progress.configure(value=progress_val))

    def reset_ui(self):
        self.btn_run.config(state="normal", text="START EXTRACTION")

# ==============================================================================
# MAIN ENTRY POINT
# ==============================================================================

if __name__ == "__main__":
    root = tk.Tk()
    app = MTCApp(root)
    root.mainloop()
